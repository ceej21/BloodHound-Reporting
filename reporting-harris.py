import sys
import argparse
import os
import json
import zipfile
import socket
from openpyxl import Workbook
#from subprocess import Popen, PIPE
import subprocess
import datetime

usage = """
$ python ./%(prog)s -d <path to .zip files> -o <output file>

Example: ./%(prog)s -d ./results/ -o output.xlsx
python reporting.py -d C:\Tools\BHTest_Data -o output
"""
parser = argparse.ArgumentParser(usage=usage)
parser.add_argument('-d', help='Path to .zip files.', dest='jsonDir', action='store')
parser.add_argument('-o', help='Output file name. Default: \'output.xlsx\'', dest='outName', action='store', default="output.xlsx")
parser.add_argument('-ip', help='Perform NSLookup.', dest='nslookup', action='store_true', default=False)
parser.add_argument('-v', help='Verbose.', dest='verbose', action='store_true', default=False)
if len(sys.argv)==1:
	parser.print_help()
	sys.exit(1)
opts = parser.parse_args()

bloodhoundJSON = {}
bloodhoundJSON["computers"] = []
bloodhoundJSON["groups"] = []
bloodhoundJSON["adminUsers"] = {}
bloodhoundJSON["administrators"] = {}
bloodhoundJSON["compAdmins"] = {}
bloodhoundJSON["computersCount"] = 0
bloodhoundJSON["adminsCount"] = 0
bloodhoundJSON["nonDA"] = 0
bloodhoundJSON["nonDAAdminUsers"] = {}
bloodhoundJSON["sessions"] = []
AD = {}
AD["collectedDomains"] = []
AD["activeDirectory"] = []

wb = Workbook()

def requestInput(question, noAnswer, yesAnswer):
	for retry in range(3):
		answer =  input("\n[?] " + question + "? [Y/n]: ") or "y"
		if answer.lower() in ("no", "n"):
			print("[*] " + noAnswer)
			return(False)
		elif answer.lower() not in ("yes", "y"):
			print("[!] %s is not a valid choice." %answer)
		else:
			print("[*] " + yesAnswer)
			return(True)
	else:
		print("\n[!] You have provided to many invalid choices. Goodbye.")
		sys.exit()

def verbosePrint(line):
	if opts.verbose:
		print(line)

if opts.outName:
	opts.outName = opts.outName.strip(".xlsx") + ".xlsx"
		
#determine if outfile exists
if os.path.isfile(opts.outName):
	if requestInput("Do you want to overwrite " + opts.outName, "Appending to existing file...", "Replacing existing file..."):
		os.remove(opts.outName)
		

def newCache():
	global bloodhoundJSON
	answer=True
	if os.path.isfile('BloodHound.json'):
		if requestInput("Do you want to use the existing BloodHound.json cache", "Creating a new cache.", "Importing existing cache..."):
			with open('BloodHound.json', "r") as f:
				bloodhoundJSON = json.load(f)
			f.close()
			answer=False
		else:
			os.remove('BloodHound.json')
	global AD
	if answer:
		if os.path.isfile('AD.json'):
			if requestInput("Do you want to use the existing AD.json cache", "Creating a new cache.", "Importing existing cache..."):
				with open('AD.json', "r") as f:

					AD = json.load(f)

				f.close()
			else:
				os.remove('AD.json')
		return True
	else:
		return False		

def queryAD(domain): 
	cmd = ['powershell.exe', '-c', 'Get-ADUser -Filter "*" -Properties EmailAddress,Manager,Enabled,Modified,LastLogonDate,DisplayName,SamAccountName,UserPrincipalName,Title,Description,Office,AccountExpirationDate,accountExpires -Server ' + domain + ' | ConvertTo-Json']
	
	r = subprocess.check_output(cmd).decode("windows-1252").lstrip()
	query = json.loads(r)
	global AD
	AD["activeDirectory"].extend(query)
	AD["collectedDomains"].append(domain.lower())


	
if newCache():
	print("[*] Collecting administrators data from BloodHound Zip Files")
	for root, dirs, files in os.walk(opts.jsonDir):
		for filename in files:
			if "zip" in filename.lower():
				cZip = zipfile.ZipFile(os.path.join(opts.jsonDir, filename))
				for i in cZip.namelist():
					if "sessions" in i:
						with cZip.open(i) as f:
							sessions = json.load(f)
						f.close()
						for session in sessions["sessions"]:
							bloodhoundJSON["sessions"].append(session)
					if "groups" in i:
						with cZip.open(i) as f:
							groups = json.load(f)
						f.close()
						for group in groups["groups"]:
							bloodhoundJSON["groups"].append(group)
							if "admin" in group["Name"].lower():
								for user in group["Members"]:
									if user["MemberType"] == "user":
										#verbosePrint(user["MemberName"] + " is in the " + group["Name"] + " group.")
										try:
											bloodhoundJSON["adminUsers"][user["MemberName"]].append(group["Name"])
										except KeyError:
											bloodhoundJSON["adminUsers"][user["MemberName"]] = []
											bloodhoundJSON["adminUsers"][user["MemberName"]].append(group["Name"])
					if "computers" in i:
						with cZip.open(i) as f:
							computers = json.load(f)
						f.close()
						for computer in computers["computers"]:
							bloodhoundJSON["computers"].append(computer)
							if computer["LocalAdmins"]:
								verbosePrint("\n    Working on " + computer["Name"] + "...")
								verbosePrint("     - Querying IP Address...")
								try:
									computer["IP"] = socket.gethostbyname(computer["Name"])
								except socket.gaierror:
									computer["IP"] = "{N/A}"
								verbosePrint("       > " + computer["IP"])
								verbosePrint("     - Listing Local Administrators...")
								for admin in computer["LocalAdmins"]:
									if admin["Type"] == "User":
										verbosePrint("       > " + admin["Name"])
										try:
											bloodhoundJSON["administrators"][admin["Name"]]["computers"][computer["Name"]] = {}
											bloodhoundJSON["administrators"][admin["Name"]]["computers"][computer["Name"]]["IP"] = computer["IP"]
										except KeyError:
											bloodhoundJSON["adminsCount"] += 1
											bloodhoundJSON["administrators"][admin["Name"]] = {}
											bloodhoundJSON["administrators"][admin["Name"]]["AD"] = {}
											bloodhoundJSON["administrators"][admin["Name"]]["computers"] = {}
											query = admin["Name"].split("@")
											if query[1].lower() not in AD["collectedDomains"]:
												verbosePrint("[*] Collecting AD Data for " + query[1] + "...")
												queryAD(query[1])
											for entry in AD["activeDirectory"]:
												if entry["SamAccountName"].lower() == query[0].lower():
													bloodhoundJSON["administrators"][admin["Name"]]["AD"]["UserName"] = entry["Name"]
													verbosePrint("           UserName: " + entry["Name"])
													bloodhoundJSON["administrators"][admin["Name"]]["AD"]["EmailAddress"] = entry["EmailAddress"]
													verbosePrint("           EmailAddress: " + str(entry["EmailAddress"]))
													bloodhoundJSON["administrators"][admin["Name"]]["AD"]["Enabled"] = entry["Enabled"]
													verbosePrint("           Enabled: " + str(entry["Enabled"]))
													bloodhoundJSON["administrators"][admin["Name"]]["AD"]["Title"] = entry["Title"]
													verbosePrint("           Title: " + str(entry["Title"]))
													bloodhoundJSON["administrators"][admin["Name"]]["AD"]["Description"] = entry["Description"]
													verbosePrint("           Description: " + str(entry["Description"]))
													bloodhoundJSON["administrators"][admin["Name"]]["AD"]["Office"] = entry["Office"]
													verbosePrint("           Office: " + str(entry["Office"]))
													if entry["Modified"]:
														bloodhoundJSON["administrators"][admin["Name"]]["AD"]["Modified"] = datetime.datetime.fromtimestamp(int(entry["Modified"].replace("/Date(","").replace(")/","")[:10])).strftime('%Y-%m-%d %I:%M:%S %p')
													else:
														bloodhoundJSON["administrators"][admin["Name"]]["AD"]["Modified"] = ""
													verbosePrint("           Modified: " + bloodhoundJSON["administrators"][admin["Name"]]["AD"]["Modified"])
													if entry["LastLogonDate"]:
														bloodhoundJSON["administrators"][admin["Name"]]["AD"]["LastLogonDate"] = datetime.datetime.fromtimestamp(int(entry["LastLogonDate"].replace("/Date(","").replace(")/","")[:10])).strftime('%Y-%m-%d %I:%M:%S %p')
													else:
														bloodhoundJSON["administrators"][admin["Name"]]["AD"]["LastLogonDate"] = ""
													verbosePrint("           LastLogonDate: " + bloodhoundJSON["administrators"][admin["Name"]]["AD"]["LastLogonDate"])
													if entry["AccountExpirationDate"]:
														bloodhoundJSON["administrators"][admin["Name"]]["AD"]["AccountExpirationDate"] = datetime.datetime.fromtimestamp(int(entry["AccountExpirationDate"].replace("/Date(","").replace(")/","")[:10])).strftime('%Y-%m-%d %I:%M:%S %p')
													else:
														bloodhoundJSON["administrators"][admin["Name"]]["AD"]["AccountExpirationDate"] = ""
													verbosePrint("           AccountExpirationDate: " + bloodhoundJSON["administrators"][admin["Name"]]["AD"]["AccountExpirationDate"])
													for i in AD["activeDirectory"]:
														if i["DistinguishedName"] == entry["Manager"]:
															bloodhoundJSON["administrators"][admin["Name"]]["AD"]["ManagerName"] = i["Name"]
															bloodhoundJSON["administrators"][admin["Name"]]["AD"]["ManagerEmail"] = i["EmailAddress"]
													if "ManagerName" not in bloodhoundJSON["administrators"][admin["Name"]]["AD"].keys():
														bloodhoundJSON["administrators"][admin["Name"]]["AD"]["ManagerName"] = "{N/A}"
													if "ManagerEmail" not in bloodhoundJSON["administrators"][admin["Name"]]["AD"].keys():
														bloodhoundJSON["administrators"][admin["Name"]]["AD"]["ManagerEmail"] = "{N/A}"
													verbosePrint("           ManagerName: " + bloodhoundJSON["administrators"][admin["Name"]]["AD"]["ManagerName"])
													try:
														verbosePrint("           ManagerEmail: " + bloodhoundJSON["administrators"][admin["Name"]]["AD"]["ManagerEmail"])
													except TypeError:
														bloodhoundJSON["administrators"][admin["Name"]]["AD"]["ManagerEmail"] = "{N/A}"
														verbosePrint("           ManagerEmail: " + bloodhoundJSON["administrators"][admin["Name"]]["AD"]["ManagerEmail"])
												'''		
												response = queryAD(query[0],query[1])
												if "ObjectNotFound" in response[0]:
													response[0] = "{N/A}"
												bloodhoundJSON["administrators"][admin["Name"]]["AD"]["UserName"] = response[0]
												verbosePrint("           UserName: " + response[0])
												bloodhoundJSON["administrators"][admin["Name"]]["AD"]["EmailAddress"] = response[1]
												verbosePrint("           EmailAddress: " + response[1])
												bloodhoundJSON["administrators"][admin["Name"]]["AD"]["Enabled"] = response[2]
												verbosePrint("           Enabled: " + response[2])
												bloodhoundJSON["administrators"][admin["Name"]]["AD"]["Modified"] = response[3]
												verbosePrint("           Modified: " + response[3])
												bloodhoundJSON["administrators"][admin["Name"]]["AD"]["LastLogonDate"] = response[4]
												verbosePrint("           LastLogonDate: " + response[4])
												bloodhoundJSON["administrators"][admin["Name"]]["AD"]["ManagerName"] = response[5]
												verbosePrint("           ManagerName: " + response[5])
												bloodhoundJSON["administrators"][admin["Name"]]["AD"]["ManagerEmail"] = response[6]
												verbosePrint("           ManagerEmail: " + response[6])
												'''
											bloodhoundJSON["administrators"][admin["Name"]]["computers"][computer["Name"]] = {}
											bloodhoundJSON["administrators"][admin["Name"]]["computers"][computer["Name"]]["IP"] = computer["IP"]
										try:
											bloodhoundJSON["compAdmins"][computer["Name"]]["admins"].append(admin["Name"])
										except KeyError:
											bloodhoundJSON["computersCount"] += 1
											bloodhoundJSON["compAdmins"][computer["Name"]] = {}
											bloodhoundJSON["compAdmins"][computer["Name"]]["IP"] = computer["IP"]
											bloodhoundJSON["compAdmins"][computer["Name"]]["admins"] = []
											bloodhoundJSON["compAdmins"][computer["Name"]]["admins"].append(admin["Name"])
	print("[*] Writing JSON data...")
	with open("BloodHound.json", 'w') as j:
		json.dump(bloodhoundJSON, j)
	j.close()
	with open("AD.json", 'w') as j:
		json.dump(AD, j)
	j.close()
	
#get count of non domain administrator accounts
for admin in bloodhoundJSON["administrators"]:
	if admin not in bloodhoundJSON["adminUsers"]:
		if "service" not in admin.lower():
			if "admin" not in admin.lower():
				if "svc" not in admin.lower():
					bloodhoundJSON["nonDA"] += 1
					bloodhoundJSON["nonDAAdminUsers"][admin] = [bloodhoundJSON["administrators"][admin]]
                    
# get local admin right users that do not belong to admin users                     
non_Admins = []        
for admin in bloodhoundJSON["administrators"]:
	if admin not in bloodhoundJSON["adminUsers"]:
		if "service" not in admin.lower():
			if "admin" not in admin.lower():
				if "svc" not in admin.lower():
					non_Admins.append(admin.lower())
                    
                                        
# get list of users who belong in admin or domain controller groups 
data = json.dumps(bloodhoundJSON["groups"])
group_Data = json.loads(data)
user_Admin_DC_Members = []
for group in group_Data:
	if "admin" in group['Name'].lower() or "domain controller" in group['Name'].lower():
		for member in group["Members"]:
			if member["MemberType"] == 'user':
				user_Admin_DC_Members.append(str(member["MemberName"]).lower())

# add non admin / domain controller group members to list 
non_Admin_DC_Group_Members = []
non_Admins = [x.lower() for x in non_Admins]
user_Admin_DC_Members = [y.lower() for y in user_Admin_DC_Members]

sA = set(user_Admin_DC_Members)
sB = set(non_Admins)

sC = sB.difference(sA)
non_Admin_DC_Group_Members = list(sC)


testA = []
#get count of non domain administrator accounts
for admin in bloodhoundJSON["administrators"]:
	if "service" not in admin.lower():
		if "admin" not in admin.lower():
			if "svc" not in admin.lower():
				adminTest = admin.split('@')[0]
				adminTest = str(adminTest) + '@'
				if adminTest.lower() in admin.lower():
					#testA.append(adminTest.lower())
					testA.append(admin.lower())

# admin users that have 'a@ admin account along with local account
check = []
for user in testA:
	try:
		if 'a@' in user:
			userSplit = user.split('a@')[0]
			userSplit = userSplit + '@'
			#print(userSplit)
			check.append(userSplit)
	except:
		continue

# admin users that have 'a@ admin account along with local account
# list of local admins
local_admins = []
for i in bloodhoundJSON["administrators"]:
	local_admins.append(i.lower())


# check if a@ account user has another account
checkAccounts = []
for user in check:
	for admin in local_admins:
		if user in admin:
			checkAccounts.append(admin.upper())


print()
print("There are " + str(bloodhoundJSON["adminsCount"]) + " accounts with admin rights to " + str(bloodhoundJSON["computersCount"]) + " machines.")
print()
print(str(bloodhoundJSON["nonDA"]) + " users out of the " + str(bloodhoundJSON["adminsCount"]) + ", are not admin/service accounts and are not in any administrator-level domain group.")
print()
print("There are " + str(len(checkAccounts)) + " local accounts with administrative rights additional to their 'a@' account")
print()
if requestInput("Do you want to generate an admin report to " + opts.outName, "Goodbye", "Writing results to " + opts.outName):
	ws0 = wb.active
	ws0.title = "Local_Administrators"
	
	ws2 = wb.create_sheet("Non_Admin_DC_Group_Members")
	ws2_cell_number = 1
	ws2['A' + str(ws2_cell_number)] = "UserName"
	ws2['B' + str(ws2_cell_number)] = "Computer"
	ws2['C' + str(ws2_cell_number)] = "IP"
	ws2['D' + str(ws2_cell_number)] = "LastLogin"
	ws2['E' + str(ws2_cell_number)] = "LastModified"
	ws2['F' + str(ws2_cell_number)] = "Title"
	ws2['G' + str(ws2_cell_number)] = "Description"
	ws2['H' + str(ws2_cell_number)] = "Office"
	ws2['I' + str(ws2_cell_number)] = "Sessions"
	ws2['J' + str(ws2_cell_number)] = "AccountExpirationDate"
	ws2_cell_number += 1
	for user in non_Admin_DC_Group_Members:
		for admin in bloodhoundJSON["administrators"]:
			if user.upper() in admin.upper():
				verbosePrint("    Writing " + admin + " to worksheet...")
				ws2['A' + str(ws2_cell_number)] = admin
				sessionsCount = 0
				for i in bloodhoundJSON["sessions"]:
					if i['UserName'] == admin:
						sessionsCount += 1

				for computer in bloodhoundJSON["administrators"][admin]["computers"]:
					ws2['A' + str(ws2_cell_number)] = admin
					ws2['B' + str(ws2_cell_number)] = computer
					try:
						ws2['C' + str(ws2_cell_number)] = \
						bloodhoundJSON["administrators"][admin]["computers"][computer]["IP"]
					except KeyError:
						ws2['C' + str(ws2_cell_number)] = "{N/A}"
					try:
						ws2['D' + str(ws2_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["LastLogonDate"]
					except KeyError:
						ws2['D' + str(ws2_cell_number)] = "{N/A}"
					try:
						ws2['E' + str(ws2_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["Modified"]
					except KeyError:
						ws2['E' + str(ws2_cell_number)] = "{N/A}"
					try:
						ws2['F' + str(ws2_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["Title"]
					except KeyError:
						ws2['F' + str(ws2_cell_number)] = "{N/A}"
					try:
						ws2['G' + str(ws2_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["Description"]
					except KeyError:
						ws2['G' + str(ws2_cell_number)] = "{N/A}"
					try:
						ws2['H' + str(ws2_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["Office"]
					except KeyError:
						ws2['H' + str(ws2_cell_number)] = "{N/A}"
					ws2['I' + str(ws2_cell_number)] = sessionsCount
					try:
						ws2['J' + str(ws2_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["AccountExpirationDate"]
					except KeyError:
						ws2['J' + str(ws2_cell_number)] = "{N/A}"
					
					ws2_cell_number += 1
	wb.save(opts.outName)
    
	ws1 = wb.create_sheet("AD_Query")
	ws0_cell_number = 1
	ws1_cell_number = 1
    
	#Local_Administrators Header Row
	ws0['A'+str(ws0_cell_number)] = "User"
	ws0['B'+str(ws0_cell_number)] = "Computer"
	ws0['C'+str(ws0_cell_number)] = "IP"
	ws0['D'+str(ws0_cell_number)] = "IT Manager"
	ws0['E'+str(ws0_cell_number)] = "NotAdmin or Admin Group Member"
	ws0['F'+str(ws0_cell_number)] = "Has Additional 'A' Account"
	ws0['G'+str(ws0_cell_number)] = "Enabled"
	ws0['H'+str(ws0_cell_number)] = "LastLogin"
	ws0['I'+str(ws0_cell_number)] = "LastModified"
	ws0['J'+str(ws0_cell_number)] = "UserName"
	ws0['K'+str(ws0_cell_number)] = "Title"
	ws0['L'+str(ws0_cell_number)] = "Description"
	ws0['M'+str(ws0_cell_number)] = "Office"
	ws0['N'+str(ws0_cell_number)] = "Sessions"
	ws0['O'+str(ws0_cell_number)] = "PowerShell"
	ws0['P'+str(ws0_cell_number)] = "AccountExpirationDate"
	ws0_cell_number += 1
	#AD_Query Header Row
	ws1['A'+str(ws1_cell_number)] = "AccountName"
	ws1['B'+str(ws1_cell_number)] = "UserName"
	ws1['C'+str(ws1_cell_number)] = "EmailAddress"
	ws1['D'+str(ws1_cell_number)] = "Enabled"
	ws1['E'+str(ws1_cell_number)] = "Modified"
	ws1['F'+str(ws1_cell_number)] = "LastLogonDate"
	ws1['G'+str(ws1_cell_number)] = "Title"
	ws1['H'+str(ws1_cell_number)] = "Office"
	ws1['I'+str(ws1_cell_number)] = "Description"
	ws1['J'+str(ws1_cell_number)] = "ManagerName"
	ws1['K'+str(ws1_cell_number)] = "ManagerEmail"
	ws1['L'+str(ws1_cell_number)] = "AccountExpirationDate"
	ws1_cell_number += 1
	
	for admin in bloodhoundJSON["administrators"]:
		verbosePrint("    Writing " + admin + " to worksheet...")
		ws1['A'+str(ws1_cell_number)] = admin
		sessionsCount = 0
		for i in bloodhoundJSON["sessions"]:
			if i['UserName'] == admin:
				sessionsCount += 1
		try:
			ws1['B'+str(ws1_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["UserName"]
		except KeyError:
			ws1['B'+str(ws1_cell_number)] = "{N/A}"
		try:
			ws1['C'+str(ws1_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["EmailAddress"]
		except KeyError:
			ws1['C'+str(ws1_cell_number)] = "{N/A}"
		try:
			ws1['D'+str(ws1_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["Enabled"]
		except KeyError:
			ws1['D'+str(ws1_cell_number)] = "{N/A}"
		try:
			ws1['E'+str(ws1_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["Modified"]
		except KeyError:
			ws1['E'+str(ws1_cell_number)] = "{N/A}"
		try:
			ws1['F'+str(ws1_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["LastLogonDate"]
		except KeyError:
			ws1['F'+str(ws1_cell_number)] = "{N/A}"
		try:
			ws1['G'+str(ws1_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["Title"]
		except KeyError:
			ws1['G'+str(ws1_cell_number)] = "{N/A}"
		try:
			ws1['H'+str(ws1_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["Description"]
		except KeyError:
			ws1['H'+str(ws1_cell_number)] = "{N/A}"
		try:
			ws1['I'+str(ws1_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["Office"]
		except KeyError:
			ws1['I'+str(ws1_cell_number)] = "{N/A}"
		try:
			ws1['J'+str(ws1_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["ManagerName"]
		except KeyError:
			ws1['J'+str(ws1_cell_number)] = "{N/A}"
		try:
			ws1['K'+str(ws1_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["ManagerEmail"]
		except KeyError:
			ws1['K'+str(ws1_cell_number)] = "{N/A}"
		try:
			ws1['L'+str(ws1_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["AccountExpirationDate"]
		except KeyError:
			ws1['L'+str(ws1_cell_number)] = "{N/A}"
		ws1_cell_number += 1

		for computer in bloodhoundJSON["administrators"][admin]["computers"]:
			ws0['A'+str(ws0_cell_number)] = admin
			ws0['B'+str(ws0_cell_number)] = computer
			try:
				ws0['C'+str(ws0_cell_number)] = bloodhoundJSON["administrators"][admin]["computers"][computer]["IP"]
			except KeyError:
				ws0['C'+str(ws0_cell_number)] = "{N/A}"
			ws0['D'+str(ws0_cell_number)] = ""

			# NOT ADMIN#

			try:
				if admin.lower() in non_Admin_DC_Group_Members:
					ws0['E'+str(ws0_cell_number)] = 'True'
				else:
					ws0['E' + str(ws0_cell_number)] = 'False'
			except KeyError:
				ws0['E' + str(ws0_cell_number)] = "{N/A}"
			# HAS 'A' Account #
			try:
				if admin.upper() in checkAccounts:
					ws0['F' + str(ws0_cell_number)] = 'True'
				else:
					ws0['F'+str(ws0_cell_number)] = 'False'
			except KeyError:
				ws0['F'+str(ws0_cell_number)] = "{N/A}"
			try:
				ws0['G'+str(ws0_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["Enabled"]
			except KeyError:
				ws0['G'+str(ws0_cell_number)] = "{N/A}"
			try:
				ws0['H'+str(ws0_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["LastLogonDate"]
			except KeyError:
				ws0['H'+str(ws0_cell_number)] = "{N/A}"
			try:
				ws0['I'+str(ws0_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["Modified"]
			except KeyError:
				ws0['I'+str(ws0_cell_number)] = "{N/A}"
			ws0['J'+str(ws0_cell_number)] = "=VLOOKUP(" + 'A'+str(ws0_cell_number) + ",AD_Query!A:B,2,0)"
			try:
				ws0['K'+str(ws0_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["Title"]
			except KeyError:
				ws0['K'+str(ws0_cell_number)] = "{N/A}"
			try:
				ws0['L'+str(ws0_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["Description"]
			except KeyError:
				ws0['L'+str(ws0_cell_number)] = "{N/A}"
			try:
				ws0['M'+str(ws0_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["Office"]
			except KeyError:
				ws0['M'+str(ws0_cell_number)] = "{N/A}"
			ws0['N'+str(ws0_cell_number)] = sessionsCount
			ID = {}
			ws0['O'+str(ws0_cell_number)] = "([ADSI]\"WinNT://" + computer + "/Administrators,group\").Remove(\"WinNT://" + admin.split('@')[1] + "/" + admin.split('@')[0] + ",user\")"
			#ws0['O'+str(ws0_cell_number)] = "cme smb " + computer + " -id " + ID[".".join(computer.split(".")[1:])] + ' -x "net localgroup administrators ' +admin + ' /delete" --exec-method smbexec'
			try:
				ws0['P'+str(ws0_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["AccountExpirationDate"]
			except KeyError:
				ws0['P'+str(ws0_cell_number)] = "{N/A}"
			
			ws0_cell_number += 1
	
        
	wb.save(opts.outName)
	
	#if requestInput("Do you want to generate a computer report", "Goodbye", "Writing results to Computer_Admins"):
	ws3 = wb.create_sheet("Computer_Admins")
	ws3_cell_number = 1
	ws3['A'+str(ws3_cell_number)] = "Computer"
	ws3['B'+str(ws3_cell_number)] = "User"
	ws3_cell_number += 1
	for computer in bloodhoundJSON["compAdmins"]:
		verbosePrint("    Writing " + computer + " to worksheet...")
		for admin in bloodhoundJSON["compAdmins"][computer]["admins"]:
			ws3['A'+str(ws3_cell_number)] = computer
			ws3['B'+str(ws3_cell_number)] = admin
			ws3_cell_number+=1
	wb.save(opts.outName)
	
	#if requestInput("Do you want to generate a sessions report", "Goodbye", "Writing results to Sessions"):
	ws4 = wb.create_sheet("Sessions")
	ws4_cell_number = 1
	ws4['A'+str(ws4_cell_number)] = "UserName"
	ws4['B'+str(ws4_cell_number)] = "ComputerName"
	ws4_cell_number += 1
	for session in bloodhoundJSON["sessions"]:
		verbosePrint("    Writing '" + session["UserName"] + " - " + session["ComputerName"] + "' to worksheet...")
		ws4['A'+str(ws4_cell_number)] = session["UserName"]
		ws4['B'+str(ws4_cell_number)] = session["ComputerName"]
		ws4_cell_number += 1
	wb.save(opts.outName)

        
#if requestInput("Do you want to generate a Report of Local Admin Users who have additional administrative accounts", "Goodbye", "Writing results to Non_A_Admin_Accounts"):
ws5 = wb.create_sheet("Non_A_Admin_Accounts")
ws5_cell_number = 1
ws5['A' + str(ws5_cell_number)] = "UserName"
ws5['B' + str(ws5_cell_number)] = "Computer"
ws5['C' + str(ws5_cell_number)] = "IP"
ws5['D' + str(ws5_cell_number)] = "LastLogin"
ws5['E' + str(ws5_cell_number)] = "LastModified"
ws5['F' + str(ws5_cell_number)] = "Title"
ws5['G' + str(ws5_cell_number)] = "Description"
ws5['H' + str(ws5_cell_number)] = "Office"
ws5['I' + str(ws5_cell_number)] = "Sessions"
ws5['J' + str(ws5_cell_number)] = "AccountExpirationDate"
ws5_cell_number += 1
for user in checkAccounts:
	for admin in bloodhoundJSON["administrators"]:
		if user.upper() in admin.upper():
			verbosePrint("    Writing " + admin + " to worksheet...")
			ws5['A' + str(ws5_cell_number)] = admin
			sessionsCount = 0
			for i in bloodhoundJSON["sessions"]:
				if i['UserName'] == admin:
					sessionsCount += 1

			for computer in bloodhoundJSON["administrators"][admin]["computers"]:
				ws5['A' + str(ws5_cell_number)] = admin
				ws5['B' + str(ws5_cell_number)] = computer
				try:
					ws5['C' + str(ws5_cell_number)] = bloodhoundJSON["administrators"][admin]["computers"][computer]["IP"]
				except KeyError:
					ws5['C' + str(ws5_cell_number)] = "{N/A}"
				try:
					ws5['D' + str(ws5_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["LastLogonDate"]
				except KeyError:
					ws5['D' + str(ws5_cell_number)] = "{N/A}"
				try:
					ws5['E' + str(ws5_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["Modified"]
				except KeyError:
					ws5['E' + str(ws5_cell_number)] = "{N/A}"
				try:
					ws5['F' + str(ws5_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["Title"]
				except KeyError:
					ws5['F' + str(ws5_cell_number)] = "{N/A}"
				try:
					ws5['G' + str(ws5_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["Description"]
				except KeyError:
					ws5['G' + str(ws5_cell_number)] = "{N/A}"
				try:
					ws5['H' + str(ws5_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["Office"]
				except KeyError:
					ws5['H' + str(ws5_cell_number)] = "{N/A}"
				ws5['I' + str(ws5_cell_number)] = sessionsCount
				try:
					ws5['J' + str(ws5_cell_number)] = bloodhoundJSON["administrators"][admin]["AD"]["AccountExpirationDate"]
				except KeyError:
					ws5['J' + str(ws5_cell_number)] = "{N/A}"
				

				ws5_cell_number += 1
wb.save(opts.outName)
#if requestInput("Would you like to generate a list of all the accounts that should be checked for necessary local admin rights", "Goodbye", "Writing results to Accounts_To_Check"):
ws6 = wb.create_sheet("Accounts_To_Check")
ws6_cell_number = 1
ws6['A' + str(ws6_cell_number)] = "UserName"

ws6_cell_number += 1
for user in non_Admin_DC_Group_Members:
	ws6['A' + str(ws6_cell_number)] = user.upper()
	ws6_cell_number +=1

ws6_cell_number +=1
for check in checkAccounts:
	ws6['A' + str(ws6_cell_number)] = check.upper()
	ws6_cell_number += 1

wb.save(opts.outName)

